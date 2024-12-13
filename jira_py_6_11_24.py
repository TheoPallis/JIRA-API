# Bugs -> 3 times each issue
# 3 minutes scraping
# 5 minutes runtime
import pandas as pd
import numpy as np
import glob
import os
from time import sleep
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from selenium.webdriver.support import expected_conditions as EC
pd.options.mode.chained_assignment = None 
import re
pd.options.display.max_rows = 100
tday = datetime.now().strftime('%d-%m-%Y')
downloads_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
import time
start_time = datetime.now()
folder = r"\\lawoffice\GSLODocuments\LegalServices_Division\01.Lawoffice_Common\BUSINESS ANALYSTS\BI JIRA\jira_app"



def export_issues(filtered_url,driver) :
    driver.get(filtered_url)

    # Wait for the "Export" button to be clickable and click it
    wait = WebDriverWait(driver, 10)
    export_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='issue-navigator-action-export-issues.ui.filter-button--trigger']")))
    export_button.click()
    sleep(10)
    buttons = driver.find_elements(By.CSS_SELECTOR, "button")
    for index, button in enumerate(buttons):
        print(button.text)
        if 'export csv (all fields)' in button.text.lower() :
            print(f"Export button found at position {index}, Element ID: {button.get_attribute('id')}")
            button.click()
            sleep(20)
            break
            

def delete_old_files(path=downloads_folder) :
     for file in os.scandir(path):
          if ("Jira Export" in file.name) or ('data_for_bi' in file.name) or ('Excel_Report_Jira' in file.name)  :
               os.remove(file)

def concatenate_files(path=downloads_folder):
    total_file = pd.concat([pd.read_csv(f) for f in glob.glob(os.path.join(path, "*fields*.csv"))], ignore_index=True)
    total_file.to_csv(os.path.join(path, "Jira_total.csv"), index=False)



# Login to Confluence and download csv report file
def jira_connect() :
    driver = webdriver.Chrome()

    new_jira = r"https://id.atlassian.com/login"
    driver.get(new_jira)
    sleep(2)
    user = driver.find_element(By.ID, "username")
    connect = driver.find_element(By.ID, "login-submit")
    user.send_keys("theodoros.pallis@sioufaslaw.gr")
    connect.click()
    password = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "password"))
    )
    password.send_keys("tops2174")
    sleep(2)
    connect.click()
    sleep(5)
    issues = driver.get("https://ambience.atlassian.net/jira/core/projects/DNY/board")
    issues_link = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='/jira/core/projects/DNY/issues']"))
    )
    # Click the element
    issues_link.click()
    
    base_url = r"https://ambience.atlassian.net/jira/core/projects/DNY/issues/?jql=project%20%3D%20%22DNY%22%20AND%20created%20%3E%3D%20%22{start}%22%20AND%20created%20%3C%3D%20%22{end}%22%20ORDER%20BY%20created%20DESC"

    # List of start and end dates for each quarter
    date_ranges = [
        ("2020-01-01", "2023-12-31"),
        ("2024-01-01", "2024-03-31"),
        ("2024-04-01", "2024-06-30"),
        ("2024-07-01", "2024-09-30"),
        ("2024-10-01", "2024-10-31"),
        ("2024-11-01", "2024-12-31")
    ]

    issue_urls = []

    for start, end in date_ranges:
        issue_urls.append(base_url.format(start=start, end=end))

    # Now `issue_urls` contains all the required URLs

    for url in issue_urls:
        issues
        export_issues(url,driver)


def processor() :
    delete_old_files()
    jira_connect()
    concatenate_files()
processor()


file = os.path.join(downloads_folder,"Jira_total.csv")
mapping_file = r"\\lawoffice\GSLODocuments\LegalServices_Division\01.Lawoffice_Common\BUSINESS ANALYSTS\BI JIRA\mapping names jira.xlsx"
mapping_df = pd.read_excel(mapping_file)
mapping_dict = mapping_df.set_index('Raw')['Cleaned'].to_dict()
dept_dict = mapping_df.set_index('Raw')['Department'].to_dict()
company_dict = mapping_df.set_index('Raw')['Company'].to_dict() 

df = pd.read_csv(file)
df = df[df['Project key'] =='DNY']
df['Company'] = df['Reporter'].map(company_dict)
df['DNY_Assignee'] = df['Assignee'].map(company_dict)
df['Department'] = df['Reporter'].map(dept_dict)
df= df.replace(mapping_dict)
df['Assignee'] = np.where((df['Assignee'].astype(str) == '0') & (df['Reporter'] == 'Βαγιάννης'), 'Βαγιάννης', 
                 np.where((df['Assignee'].astype(str) == '0') & (df['Reporter'] == 'Παπαδόπουλος'), 'Παπαδόπουλος', df['Assignee']))
def minutes(x) :
    x = str(x)
    if ';' in x :
     return (x.split(';')[-1])
    else :
        x 
def format_date(column,df) :
    df[column+"_date"] = pd.to_datetime(df[column], errors='coerce', dayfirst=True)

def extract_month(x) :
    try:
        return re.findall(r'\d{2}/\d{2}/\d{4}',x)[0].split('/')[1]
    except:
        return None
def extract_year(x) :
    try:
        return re.findall(r'\d{2}/\d{2}/\d{4}',x)[0].split('/')[2]
    except:
        return None


# Conversions
test = df
# test = test.dropna(subset='Issue key') @extra dupes1
# Replace null values with 0
test.fillna(0, inplace=True)
# 1) Μετατροπή ημερομηνίας σε date format (απαιτεί και την ρύθμιση ως ημερομηνίας εντός του bi)
format_date('Created',test)
format_date('Updated',test)
format_date('Resolved',test)
# 2) Ομαδοποίηση Status σε κατηγορίες Εκκρεμή και Μη Εκκρεμή
pending_list = ['Open', 'In Progress', 'Approved', 'For Approval', 'Specification Analysis','Reopened']
test['Εκκρεμή'] = np.where(test['Status'].isin(pending_list), "Εκκρεμή", "Μη Εκκρεμή")
# 3) Rpa and Iteration Flag
test['RPA_Flag'] = np.where(df['Summary'].str.contains('rpa', case=False, na=False), 'RPA', 'No-RPA')
test['Iteration_Flag'] = np.where(df['Labels'] == '#C.It', 'Iteration', 'No-Iteration')
 # 4) Υπολογισμός Αρχικής Εκτίμησης
test['Estimate'] = test['Original estimate'] / 3600
test['Estimate'] = test['Estimate'].astype(float).apply(lambda x : round(x,2))


# Get the non log columns
cols = ['Issue key',
       'Summary',
       'Status',
       'Created',
       'Resolved',
       'Assignee',
       'Reporter',
       'Updated_date',
       'Created_date',
       'Resolved_date',
       'Original estimate',
       'RPA_Flag',
       'Iteration_Flag',
       'Εκκρεμή',
       'Estimate',
       'Company',
       'DNY_Assignee',
       'Department']

# Get the log columns
log_cols = [f'Log Work.{i}' if i > 0 else 'Log Work' for i in range(99)]
hours_cols = ['Issue key', 'hours_assignee','before','during','after','Most_recent_update','Log_Month','Log_Year']# 'Time_before_resolved', 'Time_in_resolved','Time_after_resolved',
test.to_excel('test.xlsx')
os.startfile('test.xlsx')
def melt_df(df) :
    melted = pd.melt(df,id_vars=cols,value_vars=log_cols,value_name='Logs')
    # melted['Updated_Month'] = melted['Updated_date'].dt.month.astype(str)
    # melted['Updated_date'] = pd.to_datetime(melted['Updated_date'], errors='coerce')#
    melted['Log_Month'] = melted['Logs'].apply(extract_month).fillna("0")
    melted['Log_Year'] = melted['Logs'].apply(extract_year).fillna("0")
    # melted = melted[((melted['Log_Month'].astype(int) <= 7) & (melted['Log_Year'].astype(int) == 2024)) | (melted['Log_Year'].astype(int) == 2023)]

    melted['Log_Year'] = melted['Logs'].apply(extract_year)
    melted['Resolved_Month'] = melted['Resolved'].apply(extract_month)
    melted['Resolved_Year'] = melted['Resolved'].apply(extract_year)
    melted['minutes_assignee'] = melted['Logs'].apply(minutes)
    melted['minutes_assignee'] = melted['minutes_assignee'].fillna(0).astype(int)
    melted['hours'] = melted['minutes_assignee']  / 3600
    melted['hours'] = melted['hours'].astype(float).round(2)
    before = (melted['Log_Month'] < melted['Resolved_Month']) & (melted['Log_Year'] <= melted['Resolved_Year']) & (melted['Status'].isin(['Resolved','Closed']))
    during = (melted['Log_Month'] == melted['Resolved_Month']) & (melted['Log_Year'] == melted['Resolved_Year']) & melted['Status'].isin(['Resolved','Closed'])
    after = (melted['Log_Month'] > melted['Resolved_Month']) &  (melted['Log_Year'] >= melted['Resolved_Year']) &melted['Status'].isin(['Resolved','Closed'])
    # Calculate time spent before the resolution month, handling missing values
    melted['Time_before_resolved'] = melted[before].groupby('Issue key')['hours'].transform('sum').fillna(0)
    # Calculate time spent in the resolved month, handling missing values
    melted['Time_in_resolved'] = melted[during].groupby('Issue key')['hours'].transform('sum').fillna(0)
    # Calculate time spent after the resolution month
    melted['Time_after_resolved'] = melted[after].groupby('Issue key')['hours'].transform('sum').fillna(0)
    # Calculate the total hours spent on the issue by assignee
    melted['hours_assignee'] = melted.groupby('Issue key')['hours'].transform('sum')
    return melted


print("Unpivoting df")
# # Unpivot to get logs as rows
hours_df = melt_df(test)
print("Unpivoted df")
hours_df['before'] = hours_df.groupby('Issue key')['Time_before_resolved'].transform('min')
hours_df['after'] = hours_df.groupby('Issue key')['Time_after_resolved'].transform('min')
hours_df['during'] = hours_df.groupby('Issue key')['Time_in_resolved'].transform('min')
hours_df['Most_recent_update'] = hours_df.groupby('Issue key')['Updated_date'].transform('max')
hours_df['Most_recent_update'] = pd.to_datetime(hours_df['Most_recent_update'],format='%d/%m/%Y',errors='coerce').dt.strftime('%d/%m/%Y')
# hours_df = hours_df.drop_duplicates() @extra dupes2
hours_df.to_excel(os.path.join(folder,'hours_df.xlsx'))
# os.startfile('hours_df.xlsx')
hours_df = hours_df.drop_duplicates(subset='Issue key')
# hours_df[[hours_df['Issue key'] == 'DNY-1850']]
hours_df = hours_df[hours_cols]
test = test[cols]
merged_df = pd.merge(hours_df,test,on='Issue key',how='inner')
print("Merged df")
merged_df['Απόκλιση'] = merged_df['hours_assignee'] - merged_df['Estimate'] #processing
# merged_df['Απόκλιση'] = merged_df['Απόκλιση'].apply(lambda x : round(x,2)) 
merged_df['Εντός Εκτίμησης'] = np.where(merged_df['Απόκλιση'] <= 0, 'Εντός Εκτίμησης', 'Εκτός Εκτίμησης')#processing
merged_df['Εντός Εκτίμησης'] = np.where(merged_df['Estimate'] == 0, 'Δεν έχει δοθεί εκτίμηση', merged_df['Εντός Εκτίμησης'])#processing
merged_df['Created_Resolved_days'] = np.where(
    (merged_df['Status'] == 'Resolved') | (merged_df['Status'] == 'Closed'),
    (merged_df['Resolved_date'] - merged_df['Created_date']).dt.days,
    np.nan)
today = pd.Timestamp(datetime.today())
merged_df['Created_vs_Today_Days'] = (today - merged_df['Created_date']).dt.days
merged_df['Key'] = merged_df['Issue key'].str.strip("DNY-").astype(int)
merged_df['Created_Today'] = np.where((merged_df['Created_date'].dt.date == today.date()) & (merged_df['Created_date'].dt.year == today.year), 'Yes', 'No')
merged_df['Created_Current_Week'] = np.where((merged_df['Created_date'].dt.isocalendar().week == today.isocalendar().week) & (merged_df['Created_date'].dt.year == today.year), 'Yes', 'No')
merged_df['Created_Last_Week'] = np.where(merged_df['Created_date'].dt.isocalendar().week == (today.isocalendar().week - 1), 'Yes', 'No')
merged_df['Created_Current_Month'] = np.where((merged_df['Created_date'].dt.month == today.month) & (merged_df['Created_date'].dt.year == today.year), 'Yes', 'No')
merged_df['Resolved_Today'] = np.where((merged_df['Resolved_date'].dt.date == today.date()) & (merged_df['Resolved_date'].dt.year == today.year), 'Yes', 'No')
merged_df['Resolved_Current_Week'] = np.where((merged_df['Resolved_date'].dt.isocalendar().week == today.isocalendar().week) & (merged_df['Resolved_date'].dt.year == today.year), 'Yes', 'No')
merged_df['Resolved_Last_Week'] = np.where(merged_df['Resolved_date'].notna() & (merged_df['Resolved_date'].dt.isocalendar().week == (today.isocalendar().week - 1)), 'Yes', 'No')
merged_df['Resolved_Current_Month'] = np.where((merged_df['Resolved_date'].dt.month == today.month) & (merged_df['Resolved_date'].dt.year == today.year), 'Yes', 'No')
merged_df['Week_Resolved'] = merged_df['Resolved_date'].dt.isocalendar().week
merged_df['Week_Created'] = merged_df['Created_date'].dt.isocalendar().week

delete_old_files(folder)
merged_df.to_excel(os.path.join(folder,'data_for_bi.xlsx'))
# os.startfile(os.path.join(folder,'data_for_bi.xlsx'))




import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def format_df(file) :
    workbook = openpyxl.load_workbook(file)
    for worksheet in workbook:
        font = Font(color='FFFFFF', bold=True)
        fill = PatternFill(start_color='5552A2', end_color='5552A2', fill_type='solid')
        alignment = Alignment(horizontal='left')  # Left align
        for cell in worksheet[1]:
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
        for column in worksheet.columns:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 40      

    workbook.save(file)

clean_df = merged_df
clean_df.rename(columns={'before': 'Hours_Spent_Before_Month_of_Resolution', 
                         'during': 'Hours_Spent_In_Month_of_Resolution',
                         'after' : 'Hours_Spent_After_Month_Of_Resolution',
                         'Created_Resolved_days': 'Days_Between_Creation_and_Resolution',
                         'Created_vs_Today_Days': 'Days_Since_Creation',
                         'Απόκλιση': 'Hour_Difference_From_Estimate',
                         'hours_assignee': 'Total_Hours_Spent',                         
                         }, inplace=True)

# clean_df.drop(columns=['Original estimate','Updated_date','Created','Resolved'], inplace=True)
clean_df.sort_values(by='Key',inplace=True)
clean_df['Resolved_date'] = pd.to_datetime(clean_df['Resolved_date'],format='%d/%m/%Y',errors='coerce').dt.strftime('%d/%m/%Y')
clean_df['Created_date'] = pd.to_datetime(clean_df['Created_date'],format='%d/%m/%Y',errors='coerce').dt.strftime('%d/%m/%Y')
clean_df['Most_recent_update'] = pd.to_datetime(clean_df['Most_recent_update'],format='%d/%m/%Y',errors='coerce').dt.strftime('%d/%m/%Y')
clean_df.to_excel(os.path.join(folder,'Excel_Report_Jira.xlsx'),index=False)
format_df(os.path.join(folder,'Excel_Report_Jira.xlsx'))
os.startfile(os.path.join(folder,'Excel_Report_Jira.xlsx'))
end_time = datetime.now()
print(f"Operation complete : {start_time.strftime('%H:%M:%S')} - {end_time.strftime('%H:%M:%S')}")




# API -> can not find the log columns
# import requests
# import os
# import pandas as pd
# from requests.auth import HTTPBasicAuth
# token = r"ATATT3xFfGF0Rp2KTlX09zIQ3OOeIH3tIJjOIz33CBQ_emSapd_fvEosFokFnwFrG4rr_0JcIdL5fCd94imE20fnb5-yyAaqsDHAGy7B-fYd_2F6oUKSVT85kkrRVpEbzeD2xUSyZhAA5Vxs3DNgYrpxdX7IgTQDbnxTLGTGsOTr_18NecYzEH0=430C3820"
# jira_url  = r"https://ambience.atlassian.net/rest/api/2/search"
# username = "athanasios.koutivas@sioufaslaw.gr"  # Jira account email
# api_token = token  # API token generated from Jira account
# auth = HTTPBasicAuth(username, api_token)
# jql_query = f"project = DNY"


# start_at = 0
# max_results = 100  # Maximum allowed by Jira per request


# issues_df = pd.DataFrame()
# # Fetch issues with pagination
# while True:
#     # Parameters for the search query with pagination and selected fields
#     params = {
#         'jql': jql_query,
#         'startAt': start_at,
#         'maxResults': max_results,
#     }

   
#     # Send the request
#     response = requests.get(jira_url, params=params, auth=auth, headers={'Content-Type': 'application/json'})

#     if response.status_code == 200:
#         data = response.json()
#         issues = data.get('issues', [])

#         if not issues:
#             break  # No more issues to fetch, exit the loop

#         # Process issues into a DataFrame
#         issue_rows = []
#         for issue in issues:
#             fields = issue['fields']
#             issue_data = {
#                 'Issue Key': issue['key'],
#                 **{field: fields.get(field, '') for field in fields}
#             }
#             issue_rows.append(issue_data)
        
#         # Convert issue rows to a DataFrame and concatenate it to the main DataFrame
#         issues_df = pd.concat([issues_df, pd.DataFrame(issue_rows)], ignore_index=True)

#         # Move to the next page of results
#         start_at += max_results
#     else:
#         print(f"Failed to fetch issues, status code: {response.status_code}, response: {response.text}")
#         break

# # Save the DataFrame to an Excel file
# excel_filename = 'jira_issues.xlsx'
# issues_df.to_excel(excel_filename, index=False)

# print(f"Issues exported to {excel_filename}")
# os.startfile(excel_filename)



